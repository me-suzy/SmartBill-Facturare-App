#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pentru generarea fișierului Excel SmartBill cu date consistente
Rulează acest script și vei avea Excel-ul gata în folder!
"""

import pandas as pd
import os

from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook


from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

def format_excel_sheets(excel_path):
    """Formatează toate sheet-urile Excel pentru lizibilitate maximă"""
    print("🎨 Formatez Excel-ul pentru afișare optimă...")

    # Încarcă workbook-ul
    wb = load_workbook(excel_path)

    # Stiluri pentru formatare
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   📋 Formatez tab-ul: {sheet_name}")

        # Auto-ajustează lățimea coloanelor
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Setează lățimea cu un padding extra
            adjusted_width = min(max_length + 3, 50)  # Maxim 50 de caractere
            ws.column_dimensions[column_letter].width = adjusted_width

        # Formatează header-ul (prima linie)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Îngheață primul rând pentru scrolling
        ws.freeze_panes = "A2"

        # Adaugă filtre automate
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Salvează modificările
    wb.save(excel_path)
    print("✅ Formatarea completă!")

def create_smartbill_excel():
    print("🚀 Generez fișierul Excel SmartBill...")

    facturi_data = [
        {
            'ID': 1,
            'Numar_Factura': 'FACT-001',
            'Data_Emitere': '2025-08-01',
            'Data_Scadenta': '2025-09-01',
            'Serie_Factura': 'FACT',
            'Client_CIF': 'RO12345678',
            'Client_Nume': 'SC DEMO SRL',
            'Client_Email': 'demo@demo.ro',
            'Client_Telefon': '0721234567',
            'Client_Localitate': 'București',
            'Subtotal': 1300.00,
            'TVA': 247.00,
            'Total': 1547.00,
            'Status': 'Emisa',
            'Moneda': 'RON',
            'Curs_Schimb': 1.0,
            'Observatii': 'Factură pentru servicii IT și consultanță tehnică',
            'Termeni_Plata': '30 zile'
        },
        {
            'ID': 2,
            'Numar_Factura': 'FACT-002',
            'Data_Emitere': '2025-08-02',
            'Data_Scadenta': '2025-09-02',
            'Serie_Factura': 'FACT',
            'Client_CIF': 'RO87654321',
            'Client_Nume': 'SC TEST IMPEX',
            'Client_Email': 'office@testimex.ro',
            'Client_Telefon': '0739987654',
            'Client_Localitate': 'Cluj-Napoca',
            'Subtotal': 2800.50,
            'TVA': 532.10,
            'Total': 3332.60,
            'Status': 'Neplatita',
            'Moneda': 'RON',
            'Curs_Schimb': 1.0,
            'Observatii': 'Dezvoltare aplicații web personalizate',
            'Termeni_Plata': '30 zile'
        },
        {
            'ID': 3,
            'Numar_Factura': 'FACT-003',
            'Data_Emitere': '2025-08-03',
            'Data_Scadenta': '2025-09-03',
            'Serie_Factura': 'FACT',
            'Client_CIF': '1234567890123',
            'Client_Nume': 'Popescu Ion PFA',
            'Client_Email': 'popescu@email.ro',
            'Client_Telefon': '0743555666',
            'Client_Localitate': 'Timișoara',
            'Subtotal': 1200.00,
            'TVA': 228.00,
            'Total': 1428.00,
            'Status': 'Platita',
            'Moneda': 'RON',
            'Curs_Schimb': 1.0,
            'Observatii': 'Consultanță IT și optimizare procese',
            'Termeni_Plata': '15 zile'
        },
        {
            'ID': 4,
            'Numar_Factura': 'FACT-004',
            'Data_Emitere': '2025-08-04',
            'Data_Scadenta': '2025-09-04',
            'Serie_Factura': 'FACT',
            'Client_CIF': 'RO11223344',
            'Client_Nume': 'SC ALPHA BETA',
            'Client_Email': 'alphabeta@email.ro',
            'Client_Telefon': '0758889999',
            'Client_Localitate': 'Iași',
            'Subtotal': 3500.00,
            'TVA': 665.00,
            'Total': 4165.00,
            'Status': 'Neplatita',
            'Moneda': 'RON',
            'Curs_Schimb': 1.0,
            'Observatii': 'Audit sistem informatic complet',
            'Termeni_Plata': '30 zile'
        },
        {
            'ID': 5,
            'Numar_Factura': 'FACT-005',
            'Data_Emitere': '2025-08-05',
            'Data_Scadenta': '2025-09-05',
            'Serie_Factura': 'FACT',
            'Client_CIF': '9876543210987',
            'Client_Nume': 'Ionescu Maria PFA',
            'Client_Email': 'maria@email.ro',
            'Client_Telefon': '0766111222',
            'Client_Localitate': 'Constanța',
            'Subtotal': 890.00,
            'TVA': 169.10,
            'Total': 1059.10,
            'Status': 'Platita',
            'Moneda': 'RON',
            'Curs_Schimb': 1.0,
            'Observatii': 'Mentenanță și suport tehnic lunar',
            'Termeni_Plata': '30 zile'
        }
    ]

    # Datele pentru tab-ul CLIENTI
    # Datele pentru tab-ul CLIENTI (structura corectă)
    clienti_data = [
        {
            'ID': 1,
            'CIF': 'RO12345678',
            'Nume': 'SC DEMO SRL',
            'Client_Email': 'demo@demo.ro',
            'Client_Telefon': '0721234567',
            'Client_Localitate': 'București',
            'Client_Judet': 'București',
            'Client_Adresa': 'Strada Demo, Nr. 123, Sector 1',
            'Client_Cod_Postal': '010123',
            'Tara': 'România',
            'Website': 'www.demo.ro',
            'Persoana_Contact': 'Popescu Ion',
            'Observatii': 'Client VIP - plăți prompte',
            'Sold_Curent': 0.00,
            'Status': 'Activ'
        },
        {
            'ID': 2,
            'CIF': 'RO87654321',
            'Nume': 'SC TEST IMPEX',
            'Client_Email': 'office@testimex.ro',
            'Client_Telefon': '0739987654',
            'Client_Localitate': 'Cluj-Napoca',
            'Client_Judet': 'Cluj',
            'Client_Adresa': 'Bulevardul Testului, Nr. 456',
            'Client_Cod_Postal': '400456',
            'Tara': 'România',
            'Website': 'www.testimex.ro',
            'Persoana_Contact': 'Vasilescu Ana',
            'Observatii': 'Client nou - perioada de probă',
            'Sold_Curent': -3332.60,
            'Status': 'Activ'
        },
        {
            'ID': 3,
            'CIF': '1234567890123',
            'Nume': 'Popescu Ion PFA',
            'Client_Email': 'popescu@email.ro',
            'Client_Telefon': '0743555666',
            'Client_Localitate': 'Timișoara',
            'Client_Judet': 'Timiș',
            'Client_Adresa': 'Aleea Popescu, Nr. 789',
            'Client_Cod_Postal': '300789',
            'Tara': 'România',
            'Website': '',
            'Persoana_Contact': 'Popescu Ion',
            'Observatii': 'Colaborare pe termen lung',
            'Sold_Curent': 0.00,
            'Status': 'Activ'
        },
        {
            'ID': 4,
            'CIF': 'RO11223344',
            'Nume': 'SC ALPHA BETA',
            'Client_Email': 'alphabeta@email.ro',
            'Client_Telefon': '0758889999',
            'Client_Localitate': 'Iași',
            'Client_Judet': 'Iași',
            'Client_Adresa': 'Strada Alpha, Nr. 111',
            'Client_Cod_Postal': '700111',
            'Tara': 'România',
            'Website': 'www.alphabeta.ro',
            'Persoana_Contact': 'Ionescu Maria',
            'Observatii': 'Client corporativ - contracte mari',
            'Sold_Curent': -4165.00,
            'Status': 'Activ'
        },
        {
            'ID': 5,
            'CIF': '9876543210987',
            'Nume': 'Ionescu Maria PFA',
            'Client_Email': 'maria@email.ro',
            'Client_Telefon': '0766111222',
            'Client_Localitate': 'Constanța',
            'Client_Judet': 'Constanța',
            'Client_Adresa': 'Bulevardul Ionescu, Nr. 222',
            'Client_Cod_Postal': '900222',
            'Tara': 'România',
            'Website': '',
            'Persoana_Contact': 'Ionescu Maria',
            'Observatii': 'Plăți regulate - client de încredere',
            'Sold_Curent': 0.00,
            'Status': 'Activ'
        }
    ]

    # Datele pentru tab-ul PRODUSE (structura corectă)
    produse_data = [
        {
            'ID': 1,
            'Cod': 'SERV-001',
            'Denumire': 'Consultanță IT',
            'UM': 'ore',
            'Pret_Vanzare': 120.00,
            'Pret_Achizitie': 80.00,
            'TVA_Procent': 19,
            'Categorie': 'Servicii',
            'Stoc': 0,
            'Stoc_Minim': 0,
            'Furnizor': 'Intern',
            'Descriere': 'Servicii de consultanță tehnologică și strategică IT',
            'Cod_Bare': '',
            'Cod_Furnizor': '',
            'Greutate': 0,
            'Volum': 0,
            'Observatii': 'Serviciu pe ore - fără stoc fizic'
        },
        {
            'ID': 2,
            'Cod': 'SERV-002',
            'Denumire': 'Dezvoltare Web',
            'UM': 'proiect',
            'Pret_Vanzare': 3500.00,
            'Pret_Achizitie': 2000.00,
            'TVA_Procent': 19,
            'Categorie': 'Servicii',
            'Stoc': 0,
            'Stoc_Minim': 0,
            'Furnizor': 'Intern',
            'Descriere': 'Dezvoltare aplicații și site-uri web personalizate',
            'Cod_Bare': '',
            'Cod_Furnizor': '',
            'Greutate': 0,
            'Volum': 0,
            'Observatii': 'Serviciu complex - pricing pe proiect'
        },
        {
            'ID': 3,
            'Cod': 'PROD-001',
            'Denumire': 'Licență Software',
            'UM': 'bucată',
            'Pret_Vanzare': 250.00,
            'Pret_Achizitie': 150.00,
            'TVA_Procent': 19,
            'Categorie': 'Software',
            'Stoc': 50,
            'Stoc_Minim': 10,
            'Furnizor': 'Microsoft Romania',
            'Descriere': 'Licențe software pentru aplicații business',
            'Cod_Bare': '1234567890123',
            'Cod_Furnizor': 'MS-LIC-001',
            'Greutate': 0,
            'Volum': 0,
            'Observatii': 'Licențe digitale - stoc virtual'
        },
        {
            'ID': 4,
            'Cod': 'PROD-002',
            'Denumire': 'Laptop Business',
            'UM': 'bucată',
            'Pret_Vanzare': 3800.00,
            'Pret_Achizitie': 2500.00,
            'TVA_Procent': 19,
            'Categorie': 'Hardware',
            'Stoc': 8,
            'Stoc_Minim': 3,
            'Furnizor': 'Dell Romania',
            'Descriere': 'Laptop-uri pentru business și dezvoltare',
            'Cod_Bare': '6675210987',
            'Cod_Furnizor': 'DELL-LAT-5520',
            'Greutate': 1.8,
            'Volum': 2.5,
            'Observatii': 'Configurat și optimizat pentru development'
        },
        {
            'ID': 5,
            'Cod': 'SERV-003',
            'Denumire': 'Suport Tehnic',
            'UM': 'ore',
            'Pret_Vanzare': 95.00,
            'Pret_Achizitie': 60.00,
            'TVA_Procent': 19,
            'Categorie': 'Servicii',
            'Stoc': 0,
            'Stoc_Minim': 0,
            'Furnizor': 'Intern',
            'Descriere': 'Suport tehnic și mentenanță sisteme IT',
            'Cod_Bare': '',
            'Cod_Furnizor': '',
            'Greutate': 0,
            'Volum': 0,
            'Observatii': 'Disponibil 24/7 pentru clienți VIP'
        }
    ]

    # Datele pentru tab-ul BONURI_FISCALE (structura corectă)
    bonuri_fiscale_data = [
        {
            'ID': 1,
            'Numar_Bon': 'BON-001',
            'Data_Emitere': '2025-08-01',
            'Ora_Emitere': '09:15:30',
            'Serie_Bon': 'DY123456',
            'Numar_Casa': '001',
            'Operator': 'Popescu Maria',
            'Client_Nume': 'SC DEMO SRL',
            'Client_Email': 'demo@demo.ro',
            'Client_Telefon': '0721234567',
            'Subtotal': 450.00,
            'TVA': 85.50,
            'Total': 535.50,
            'Metoda_Plata': 'Card',
            'Suma_Primita': 535.50,
            'Rest': 0.00,
            'Observatii': 'Bon fiscal pentru vânzare directă'
        },
        {
            'ID': 2,
            'Numar_Bon': 'BON-002',
            'Data_Emitere': '2025-08-01',
            'Ora_Emitere': '11:22:45',
            'Serie_Bon': 'DY123456',
            'Numar_Casa': '001',
            'Operator': 'Ionescu Ana',
            'Client_Nume': 'Popescu Ion PFA',
            'Client_Email': 'popescu@email.ro',
            'Client_Telefon': '0743555666',
            'Subtotal': 280.00,
            'TVA': 53.20,
            'Total': 333.20,
            'Metoda_Plata': 'Numerar',
            'Suma_Primita': 350.00,
            'Rest': 16.80,
            'Observatii': 'Plată cash - marfă livrată'
        },
        {
            'ID': 3,
            'Numar_Bon': 'BON-003',
            'Data_Emitere': '2025-08-02',
            'Ora_Emitere': '14:30:12',
            'Serie_Bon': 'DY789012',
            'Numar_Casa': '002',
            'Operator': 'Gheorghe Mihai',
            'Client_Nume': 'SC TEST IMPEX',
            'Client_Email': 'office@testimex.ro',
            'Client_Telefon': '0739987654',
            'Subtotal': 750.00,
            'TVA': 142.50,
            'Total': 892.50,
            'Metoda_Plata': 'Transfer bancar',
            'Suma_Primita': 892.50,
            'Rest': 0.00,
            'Observatii': 'Transfer instant - client corporativ'
        }
    ]

    # Datele pentru tab-ul FACTURI_STORNO (structura corectă)
    facturi_storno_data = [
        {
            'ID': 1,
            'Numar_Factura_Originala': 'FACT-002',
            'Data_Emitere': '2025-08-05',
            'Serie_Factura': 'STORNO',
            'Client_CIF': 'RO87654321',
            'Client_Nume': 'SC TEST IMPEX',
            'Client_Email': 'office@testimex.ro',
            'Client_Telefon': '0739987654',
            'Client_Localitate': 'Cluj-Napoca',
            'Subtotal_Original': 2800.50,
            'TVA_Original': 532.10,
            'Total_Original': 3332.60,
            'Subtotal_Storno': -500.00,
            'TVA_Storno': -95.00,
            'Total_Storno': -595.00,
            'Motiv_Storno': 'Eroare în cantitatea facturată',
            'Status': 'Aprobat',
            'Observatii': 'Storno parțial - cantitate greșită pe un produs'
        },
        {
            'ID': 2,
            'Numar_Factura_Originala': 'FACT-004',
            'Data_Emitere': '2025-08-07',
            'Serie_Factura': 'STORNO',
            'Client_CIF': 'RO11223344',
            'Client_Nume': 'SC ALPHA BETA',
            'Client_Email': 'alphabeta@email.ro',
            'Client_Telefon': '0758889999',
            'Client_Localitate': 'Iași',
            'Subtotal_Original': 3500.00,
            'TVA_Original': 665.00,
            'Total_Original': 4165.00,
            'Subtotal_Storno': -300.00,
            'TVA_Storno': -57.00,
            'Total_Storno': -357.00,
            'Motiv_Storno': 'Preț incorect aplicat',
            'Status': 'In asteptare',
            'Observatii': 'Așteaptă aprobarea managerului pentru reducerea prețului'
        }
    ]

    # Datele pentru tab-ul CONFIGURARE
    # Datele pentru tab-ul CONFIGURARE (structura corectă)
    configurare_data = [
        {
            'Nume_Companie': 'TIP B. SRL',
            'CIF': 'RO40123456',
            'Adresa': 'Str. Principală, Nr. 100, București, Sector 1',
            'Telefon': '021.123.4567',
            'Email': 'ioan.fantanaru@gmail.com',
            'Website': 'www.tipb.ro',
            'Cont_Bancar': 'RO49 AAAA 1B31 0075 9384 0000',
            'Banca': 'Banca Transilvania',
            'TVA': '19',
            'Moneda': 'RON'
        },

        {
            'ID': 2,
            'Categorie': 'Companie',
            'Cheie': 'cif_companie',
            'Valoare': 'RO40123456',
            'Descriere': 'Codul de identificare fiscală',
            'Tip_Date': 'text',
            'Modificabil': 'Nu',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 3,
            'Categorie': 'Companie',
            'Cheie': 'adresa_companie',
            'Valoare': 'Str. Principală, Nr. 100, București, Sector 1',
            'Descriere': 'Adresa sediului social',
            'Tip_Date': 'text',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 4,
            'Categorie': 'Facturare',
            'Cheie': 'tva_default',
            'Valoare': '19',
            'Descriere': 'Cota TVA implicită pentru produse (%)',
            'Tip_Date': 'numeric',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 5,
            'Categorie': 'Facturare',
            'Cheie': 'moneda_default',
            'Valoare': 'RON',
            'Descriere': 'Moneda implicită pentru facturi',
            'Tip_Date': 'text',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 6,
            'Categorie': 'Facturare',
            'Cheie': 'numar_factura_prefix',
            'Valoare': 'FACT-',
            'Descriere': 'Prefixul pentru numerele de factură',
            'Tip_Date': 'text',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 7,
            'Categorie': 'Facturare',
            'Cheie': 'termeni_plata_default',
            'Valoare': '30',
            'Descriere': 'Termenii de plată implicite (zile)',
            'Tip_Date': 'numeric',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 8,
            'Categorie': 'Email',
            'Cheie': 'smtp_server',
            'Valoare': 'smtp.gmail.com',
            'Descriere': 'Serverul SMTP pentru trimiterea emailurilor',
            'Tip_Date': 'text',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 9,
            'Categorie': 'Email',
            'Cheie': 'email_sender',
            'Valoare': 'ioan.fantanaru@gmail.com',
            'Descriere': 'Adresa email pentru trimiterea facturilor',
            'Tip_Date': 'email',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 10,
            'Categorie': 'Sistem',
            'Cheie': 'versiune_aplicatie',
            'Valoare': '1.0.0',
            'Descriere': 'Versiunea curentă a aplicației SmartBill',
            'Tip_Date': 'text',
            'Modificabil': 'Nu',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Sistem'
        },
        {
            'ID': 11,
            'Categorie': 'Backup',
            'Cheie': 'backup_automat',
            'Valoare': 'true',
            'Descriere': 'Activează backup-ul automat al bazei de date',
            'Tip_Date': 'boolean',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        },
        {
            'ID': 12,
            'Categorie': 'Backup',
            'Cheie': 'interval_backup',
            'Valoare': '24',
            'Descriere': 'Intervalul pentru backup automat (ore)',
            'Tip_Date': 'numeric',
            'Modificabil': 'Da',
            'Data_Actualizare': '2025-01-15',
            'Utilizator': 'Administrator'
        }
    ]

    # Creează DataFrame-urile
    df_facturi = pd.DataFrame(facturi_data)
    df_clienti = pd.DataFrame(clienti_data)
    df_produse = pd.DataFrame(produse_data)

    # Creează DataFrame-urile pentru noile tab-uri
    df_bonuri_fiscale = pd.DataFrame(bonuri_fiscale_data)
    df_facturi_storno = pd.DataFrame(facturi_storno_data)
    df_configurare = pd.DataFrame(configurare_data)

    # Creează directorul dacă nu există
    os.makedirs(r'd:\SmartBillLikeApp\database', exist_ok=True)

    # Calea către fișierul Excel
    excel_path = r'd:\SmartBillLikeApp\database\smartbill_database.xlsx'

    # Scrie în Excel cu formatare avansată
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Configurează opțiunile pentru fiecare sheet
        sheets_config = {
            'Facturi': df_facturi,
            'Bonuri_Fiscale': df_bonuri_fiscale,
            'Facturi_Storno': df_facturi_storno,
            'Configurare': df_configurare,
            'Produse': df_produse,
            'Clienti': df_clienti
        }

        for sheet_name, dataframe in sheets_config.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)

            # Accesează worksheet-ul pentru formatare
            worksheet = writer.sheets[sheet_name]

            # Auto-ajustează lățimea coloanelor
            for idx, col in enumerate(dataframe.columns):
                # Calculează lățimea optimă
                max_len = max(
                    dataframe[col].astype(str).str.len().max(),  # Cel mai lung text din coloană
                    len(str(col))  # Lungimea numelui coloanei
                )
                # Adaugă padding și limitează la maxim 50
                adjusted_width = min(max_len + 5, 50)
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width

            # Formatează header-ul
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Îngheață primul rând
            worksheet.freeze_panes = "A2"

            # Adaugă filtre automate
            if len(dataframe) > 0:
                worksheet.auto_filter.ref = f"A1:{get_column_letter(len(dataframe.columns))}{len(dataframe) + 1}"

    print(f"✅ Fișierul Excel a fost creat cu succes!")
    print(f"📁 Locația: {excel_path}")
    print(f"📊 Conținut formatat profesional:")
    print(f"   - Tab 'Facturi': {len(df_facturi)} înregistrări")
    print(f"   - Tab 'Bonuri_Fiscale': {len(df_bonuri_fiscale)} înregistrări")
    print(f"   - Tab 'Facturi_Storno': {len(df_facturi_storno)} înregistrări")
    print(f"   - Tab 'Configurare': {len(df_configurare)} înregistrări")
    print(f"   - Tab 'Produse': {len(df_produse)} înregistrări")
    print(f"   - Tab 'Clienti': {len(df_clienti)} înregistrări")

    print(f"\n✨ Funcții de formatare aplicate:")
    print(f"   ✅ Coloane auto-ajustate la conținut")
    print(f"   ✅ Header-uri cu fundal albastru și text alb")
    print(f"   ✅ Primul rând înghețat pentru scrolling")
    print(f"   ✅ Filtre automate pe toate coloanele")
    print(f"   ✅ Spațiere optimă între coloane")

    print(f"\n🚀 Excel-ul este gata și formatat profesional!")

    # Verifică consistența datelor
    facturi_cifs = set(df_facturi['Client_CIF'].unique())
    clienti_cifs = set(df_clienti['CIF'].unique())

    missing_clients = facturi_cifs - clienti_cifs
    if missing_clients:
        print(f"❌ CIF-uri din facturi care nu sunt în clienți: {missing_clients}")
    else:
        print(f"✅ Toate CIF-urile din facturi există în tab-ul clienți")

    # Verifică numele clienților
    consistente = True
    for _, factura in df_facturi.iterrows():
        client_row = df_clienti[df_clienti['CIF'] == factura['Client_CIF']]
        if not client_row.empty:
            if client_row.iloc[0]['Nume'] != factura['Client_Nume']:
                print(f"❌ Discrepanță nume pentru CIF {factura['Client_CIF']}: '{factura['Client_Nume']}' vs '{client_row.iloc[0]['Nume']}'")
                consistente = False

    if consistente:
        print(f"✅ Toate numele clienților sunt consistente între tab-uri")

    print(f"\n🚀 Excel-ul este gata de folosit!")
    print(f"💡 Acum poți rula aplicația SmartBill și testa căutarea facturilor.")

if __name__ == "__main__":
    try:
        create_smartbill_excel()
    except ImportError:
        print("❌ Eroare: pandas nu este instalat.")
        print("💡 Rulează: pip install pandas openpyxl")
        print("   Apoi rulează din nou acest script.")
    except Exception as e:
        print(f"❌ Eroare: {e}")
        print("💡 Asigură-te că ai permisiuni de scriere în directorul specificat.")